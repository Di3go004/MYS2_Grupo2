"""
Simio Aerospace - Fase 1 - Analisis de escenarios (Joaco)
============================================================
Recalcula desde los datos crudos (Simio Aerospace Data.xlsx) los 4 calculos de
capacidad -para verificar de forma independiente lo que ya entrego Madelyn- y
construye el analisis de escenarios que le corresponde a Joaco:

    "Comparacion de la carga de trabajo correspondiente al lote inicial de 75
    aviones" (60% SA101 / 40% SA102).

Ademas compara esa carga contra la capacidad teorica disponible (turno / dia /
semana) y contra la duracion de calendario que impone la politica actual de
ciclo fijo de 4 dias, para alimentar el analisis de cuellos de botella.

Entradas:
    - Simio Aerospace Data.xlsx           (datos crudos, hojas WorkCell1..5)
Salidas:
    - Joaco_Fase1/analisis/Analisis_Escenarios_Lote75.xlsx
    - Joaco_Fase1/analisis/resumen_escenarios.json  (para las graficas)
"""
import json
import re
import math
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[2]
RAW_DATA = ROOT / "Simio Aerospace Data.xlsx"
OUT_DIR = Path(__file__).resolve().parent
OUT_XLSX = OUT_DIR / "Analisis_Escenarios_Lote75.xlsx"
OUT_JSON = OUT_DIR / "resumen_escenarios.json"

CELDAS = ["WorkCell1", "WorkCell2", "WorkCell3", "WorkCell4", "WorkCell5"]

# ---- Parametros del caso (enunciado Fase 1) --------------------------------
LOTE_TOTAL = 75
PCT_SA101 = 0.60
PCT_SA102 = 0.40
N_SA101 = round(LOTE_TOTAL * PCT_SA101)          # 45
N_SA102 = LOTE_TOTAL - N_SA101                    # 30
CICLO_DIAS = 4                                    # dias por avion (politica actual)
HORAS_EFECTIVAS_POR_CICLO = None                  # se calcula abajo (depende de HORAS_DIA)
MECANICOS_POR_CELDA = 8
HORAS_TURNO1 = 4 + 4          # 6:00-10:00 + 10:30-14:30
HORAS_TURNO2 = 4 + 3          # 14:30-18:30 + 19:00-22:00
HORAS_DIA = HORAS_TURNO1 + HORAS_TURNO2           # 15 h efectivas/dia por mecanico
DIAS_SEMANA = 5
HH_DIA_CELDA = MECANICOS_POR_CELDA * HORAS_DIA            # 120 HH/dia
HH_SEMANA_CELDA = HH_DIA_CELDA * DIAS_SEMANA               # 600 HH/semana
HORAS_EFECTIVAS_POR_CICLO = HORAS_DIA * CICLO_DIAS         # 60 h efectivas por ciclo de 4 dias

TRI_RE = re.compile(r"Random\.Triangular\(\s*([\d.]+)\s*,\s*([\d.]+)\s*,\s*([\d.]+)\s*\)")


def parse_triangular(expr):
    """Devuelve (a, m, b, media, varianza, desv_est) de un texto
    'Random.Triangular(a,m,b)'."""
    match = TRI_RE.search(str(expr))
    if not match:
        raise ValueError(f"No se pudo interpretar la expresion triangular: {expr!r}")
    a, m, b = (float(x) for x in match.groups())
    media = (a + m + b) / 3
    varianza = (a**2 + m**2 + b**2 - a * m - a * b - m * b) / 18
    desv = math.sqrt(varianza)
    return a, m, b, media, varianza, desv


def load_workcell(ws):
    """Lee una hoja WorkCellN y regresa una lista de tareas con sus metricas."""
    tasks = []
    rows = list(ws.iter_rows(values_only=True))
    for row in rows[2:]:  # las 2 primeras filas son encabezado
        if row[0] is None or row[1] is None:
            continue
        seq, task, lh1, lh2, mech = row[0], row[1], row[2], row[3], row[4]
        a1, m1, b1, mean1, var1, sd1 = parse_triangular(lh1)
        a2, m2, b2, mean2, var2, sd2 = parse_triangular(lh2)
        tasks.append({
            "seq": int(seq),
            "task": str(task).strip(),
            "mechanics": int(mech),
            "sa101_mean": mean1, "sa101_sd": sd1,
            "sa102_mean": mean2, "sa102_sd": sd2,
            "hh101": mean1 * int(mech),
            "hh102": mean2 * int(mech),
        })
    return tasks


def analyze_celda(tasks):
    """Aplica los calculos 1-3 (carga, horas-hombre, duracion por grupo paralelo)."""
    carga101 = sum(t["sa101_mean"] for t in tasks)
    carga102 = sum(t["sa102_mean"] for t in tasks)
    hh101 = sum(t["hh101"] for t in tasks)
    hh102 = sum(t["hh102"] for t in tasks)

    seqs = sorted(set(t["seq"] for t in tasks))
    grupos = []
    for s in seqs:
        grupo = [t for t in tasks if t["seq"] == s]
        mech_grupo = sum(t["mechanics"] for t in grupo)
        dur101 = max(t["sa101_mean"] for t in grupo)
        dur102 = max(t["sa102_mean"] for t in grupo)
        grupos.append({
            "seq": s, "n_tareas": len(grupo), "mecanicos_simultaneos": mech_grupo,
            "excede_8": mech_grupo > MECANICOS_POR_CELDA,
            "duracion101": dur101, "duracion102": dur102,
        })
    duracion_total101 = sum(g["duracion101"] for g in grupos)
    duracion_total102 = sum(g["duracion102"] for g in grupos)
    pico_mecanicos = max(g["mecanicos_simultaneos"] for g in grupos)

    return {
        "n_tareas": len(tasks),
        "carga101": carga101, "carga102": carga102,
        "hh101": hh101, "hh102": hh102,
        "duracion_total101": duracion_total101, "duracion_total102": duracion_total102,
        "pico_mecanicos": pico_mecanicos,
        "grupos": grupos,
        "tasks": tasks,
    }


def main():
    wb_raw = openpyxl.load_workbook(RAW_DATA, data_only=True)
    resultados = {}
    for celda in CELDAS:
        tasks = load_workcell(wb_raw[celda])
        resultados[celda] = analyze_celda(tasks)

    # ---- Escenario: lote de 75 aviones (45 SA101 / 30 SA102) --------------
    dias_calendario_lote = (LOTE_TOTAL - 1) * CICLO_DIAS + len(CELDAS) * CICLO_DIAS
    semanas_calendario_lote = dias_calendario_lote / DIAS_SEMANA

    escenario = {}
    total_hh_lote = 0
    for celda, r in resultados.items():
        hh_lote = r["hh101"] * N_SA101 + r["hh102"] * N_SA102
        dias_mano_obra = hh_lote / HH_DIA_CELDA
        semanas_mano_obra = dias_mano_obra / DIAS_SEMANA
        utilizacion_pct = (hh_lote / (dias_calendario_lote * HH_DIA_CELDA)) * 100

        # --- Riesgo de trabajo desplazado por ciclo: compara la duracion real
        # de un avion en la celda (Calculo 3, asumiendo 8 mecanicos) contra las
        # horas efectivas disponibles en una ventana de 4 dias (60 h). Si la
        # duracion excede la ventana, ese excedente es trabajo que la politica
        # actual "desplaza" a la siguiente celda.
        excedente101 = max(0, r["duracion_total101"] - HORAS_EFECTIVAS_POR_CICLO)
        excedente102 = max(0, r["duracion_total102"] - HORAS_EFECTIVAS_POR_CICLO)
        pct_desplazado101 = excedente101 / r["duracion_total101"] * 100 if r["duracion_total101"] else 0
        pct_desplazado102 = excedente102 / r["duracion_total102"] * 100 if r["duracion_total102"] else 0

        escenario[celda] = {
            "hh_avion_sa101": r["hh101"],
            "hh_avion_sa102": r["hh102"],
            "hh_lote": hh_lote,
            "pct_del_total_lote": None,  # se llena despues
            "dias_solo_mano_obra": dias_mano_obra,
            "semanas_solo_mano_obra": semanas_mano_obra,
            "utilizacion_bajo_ciclo_fijo_pct": utilizacion_pct,
            "pico_mecanicos_simultaneos": r["pico_mecanicos"],
            "pct_uso_pico_mecanicos": r["pico_mecanicos"] / MECANICOS_POR_CELDA * 100,
            "duracion_avion_sa101_h": r["duracion_total101"],
            "duracion_avion_sa102_h": r["duracion_total102"],
            "excedente_sa101_h": excedente101,
            "excedente_sa102_h": excedente102,
            "pct_trabajo_desplazado_sa101": pct_desplazado101,
            "pct_trabajo_desplazado_sa102": pct_desplazado102,
        }
        total_hh_lote += hh_lote

    for celda in escenario:
        escenario[celda]["pct_del_total_lote"] = escenario[celda]["hh_lote"] / total_hh_lote * 100

    resumen = {
        "parametros": {
            "lote_total": LOTE_TOTAL, "n_sa101": N_SA101, "n_sa102": N_SA102,
            "ciclo_dias": CICLO_DIAS, "hh_dia_celda": HH_DIA_CELDA,
            "hh_semana_celda": HH_SEMANA_CELDA,
            "dias_calendario_lote_ciclo_fijo": dias_calendario_lote,
            "semanas_calendario_lote_ciclo_fijo": semanas_calendario_lote,
        },
        "por_celda_base": {
            celda: {
                "n_tareas": r["n_tareas"],
                "carga101": r["carga101"], "carga102": r["carga102"],
                "hh101": r["hh101"], "hh102": r["hh102"],
                "duracion_total101": r["duracion_total101"],
                "duracion_total102": r["duracion_total102"],
                "pico_mecanicos": r["pico_mecanicos"],
                "grupos": r["grupos"],
            } for celda, r in resultados.items()
        },
        "escenario_lote_75": escenario,
        "total_hh_lote_75": total_hh_lote,
    }

    OUT_JSON.write_text(json.dumps(resumen, indent=2, ensure_ascii=False), encoding="utf-8")
    print(f"JSON escrito en {OUT_JSON}")

    # =========================================================================
    # Libro de Excel editable (fuente reproducible, formulas visibles como valores
    # + notas de formula para que cualquiera pueda auditar el calculo)
    # =========================================================================
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    bold = Font(bold=True)
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(bold=True, color="FFFFFF")
    warn_fill = PatternFill("solid", fgColor="FFC7CE")
    ok_fill = PatternFill("solid", fgColor="C6EFCE")

    # ---- Hoja 1: Parametros -------------------------------------------------
    ws = wb.create_sheet("Parametros")
    ws.append(["Parametro", "Valor", "Fuente / justificacion"])
    for c in ws[1]:
        c.font = header_font; c.fill = header_fill
    filas_param = [
        ("Lote total de aviones", LOTE_TOTAL, "Enunciado Fase 1"),
        ("% SA101", f"{PCT_SA101:.0%}", "Enunciado Fase 1"),
        ("% SA102", f"{PCT_SA102:.0%}", "Enunciado Fase 1"),
        ("Aviones SA101 en el lote", N_SA101, "75 x 60%"),
        ("Aviones SA102 en el lote", N_SA102, "75 x 40%"),
        ("Ciclo fijo actual (dias/avion)", CICLO_DIAS, "Enunciado Fase 1"),
        ("Mecanicos por celda (por turno)", MECANICOS_POR_CELDA, "Enunciado Fase 1"),
        ("Horas efectivas turno 1", HORAS_TURNO1, "6:00-10:00 + 10:30-14:30"),
        ("Horas efectivas turno 2", HORAS_TURNO2, "14:30-18:30 + 19:00-22:00"),
        ("Horas efectivas por dia (2 turnos)", HORAS_DIA, "Turno1 + Turno2"),
        ("Dias laborales por semana", DIAS_SEMANA, "Lunes a viernes"),
        ("Capacidad HH/dia por celda", HH_DIA_CELDA, "8 mecanicos x 15 h"),
        ("Capacidad HH/semana por celda", HH_SEMANA_CELDA, "120 HH/dia x 5 dias"),
        ("Duracion de calendario del lote bajo ciclo fijo (dias)", dias_calendario_lote,
         "(75-1)*4 + 5 celdas*4 = tiempo entre 1er avion entrando y ultimo saliendo de celda 5"),
        ("Duracion de calendario del lote bajo ciclo fijo (semanas)", round(semanas_calendario_lote, 2), "dias / 5"),
        ("Horas efectivas disponibles por ciclo (4 dias)", HORAS_EFECTIVAS_POR_CICLO,
         "15 h efectivas/dia x 4 dias -- ventana real de trabajo antes de que el avion deba avanzar"),
    ]
    for f in filas_param:
        ws.append(f)
    ws.column_dimensions["A"].width = 52
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 55

    # ---- Hoja 2: Escenario lote de 75 aviones ------------------------------
    ws = wb.create_sheet("Escenario_Lote75")
    headers = [
        "Celda", "HH esperadas / avion SA101", "HH esperadas / avion SA102",
        "HH totales del lote (45 SA101 + 30 SA102)", "% del total del lote",
        "Dias necesarios (solo mano de obra, sin ciclo fijo)",
        "Semanas necesarias (solo mano de obra)",
        f"Utilizacion bajo ciclo fijo actual ({dias_calendario_lote} dias) %",
        "Pico de mecanicos simultaneos (Calculo 3)", "% de uso del pico vs 8 disponibles",
    ]
    ws.append(headers)
    for c in ws[1]:
        c.font = header_font; c.fill = header_fill
        c.alignment = Alignment(wrap_text=True, vertical="center")
    ws.row_dimensions[1].height = 45

    max_hh_lote = max(v["hh_lote"] for v in escenario.values())
    for celda in CELDAS:
        e = escenario[celda]
        row = [
            celda.replace("WorkCell", "Celda "),
            round(e["hh_avion_sa101"], 2), round(e["hh_avion_sa102"], 2),
            round(e["hh_lote"], 2), round(e["pct_del_total_lote"], 2),
            round(e["dias_solo_mano_obra"], 2), round(e["semanas_solo_mano_obra"], 2),
            round(e["utilizacion_bajo_ciclo_fijo_pct"], 2),
            e["pico_mecanicos_simultaneos"], round(e["pct_uso_pico_mecanicos"], 1),
        ]
        ws.append(row)
        r = ws.max_row
        if e["hh_lote"] == max_hh_lote:
            for col in range(1, len(headers) + 1):
                ws.cell(row=r, column=col).fill = warn_fill
        if e["pico_mecanicos_simultaneos"] >= MECANICOS_POR_CELDA:
            ws.cell(row=r, column=9).fill = warn_fill
            ws.cell(row=r, column=10).fill = warn_fill

    ws.append(["TOTAL", "", "", round(total_hh_lote, 2), 100.0, "", "", "", "", ""])
    for c in ws[ws.max_row]:
        c.font = bold
    for i, w in enumerate([14, 16, 16, 18, 14, 16, 16, 16, 16, 16], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # ---- Hoja 2b: Riesgo de trabajo desplazado por ciclo fijo (hallazgo clave)
    ws = wb.create_sheet("Riesgo_Trabajo_Desplazado")
    ws.append([f"Comparacion: duracion real de 1 avion en la celda (Calculo 3, 8 mecanicos) "
               f"vs. horas efectivas disponibles en un ciclo de {CICLO_DIAS} dias "
               f"({HORAS_EFECTIVAS_POR_CICLO} h)"])
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=8)
    ws["A1"].font = bold
    ws.append(["Celda", "Duracion avion SA101 (h)", "Excede ventana de 60h?",
               "Trabajo desplazado SA101 (h)", "% desplazado SA101",
               "Duracion avion SA102 (h)", "Trabajo desplazado SA102 (h)", "% desplazado SA102"])
    for c in ws[2]:
        c.font = header_font; c.fill = header_fill
        c.alignment = Alignment(wrap_text=True, vertical="center")
    ws.row_dimensions[2].height = 45
    for celda in CELDAS:
        e = escenario[celda]
        ws.append([
            celda.replace("WorkCell", "Celda "),
            round(e["duracion_avion_sa101_h"], 2),
            "SI" if e["excedente_sa101_h"] > 0 else "no",
            round(e["excedente_sa101_h"], 2), round(e["pct_trabajo_desplazado_sa101"], 1),
            round(e["duracion_avion_sa102_h"], 2),
            round(e["excedente_sa102_h"], 2), round(e["pct_trabajo_desplazado_sa102"], 1),
        ])
        if e["excedente_sa101_h"] > 0 or e["excedente_sa102_h"] > 0:
            for col in range(1, 9):
                ws.cell(row=ws.max_row, column=col).fill = warn_fill
        else:
            for col in range(1, 9):
                ws.cell(row=ws.max_row, column=col).fill = ok_fill
    for i, w in enumerate([14, 20, 18, 20, 16, 20, 20, 16], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.append([])
    ws.append(["Lectura: si 'Trabajo desplazado (h)' > 0, bajo la politica actual de ciclo fijo "
               "esa celda no alcanza a terminar el avion dentro de sus 4 dias y empuja ese trabajo "
               "pendiente a la siguiente celda -- exactamente el mecanismo de 'trabajo desplazado' "
               "que describe el caso."])
    ws.merge_cells(start_row=ws.max_row, start_column=1, end_row=ws.max_row, end_column=8)

    # ---- Hoja 3: Detalle por celda + grupos de secuencia -------------------
    ws = wb.create_sheet("Detalle_Grupos_Secuencia")
    ws.append(["Celda", "Secuencia", "# Tareas en paralelo", "Mecanicos simultaneos",
               "Excede 8 disponibles", "Duracion grupo SA101 (h)", "Duracion grupo SA102 (h)"])
    for c in ws[1]:
        c.font = header_font; c.fill = header_fill
    for celda, r in resultados.items():
        for g in r["grupos"]:
            ws.append([
                celda.replace("WorkCell", "Celda "), g["seq"], g["n_tareas"],
                g["mecanicos_simultaneos"], "SI" if g["excede_8"] else "no",
                round(g["duracion101"], 2), round(g["duracion102"], 2),
            ])
            if g["excede_8"] or g["mecanicos_simultaneos"] == MECANICOS_POR_CELDA:
                fill = warn_fill if g["excede_8"] else ok_fill
                for col in range(1, 8):
                    ws.cell(row=ws.max_row, column=col).fill = fill
    for i, w in enumerate([14, 12, 18, 18, 16, 20, 20], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # ---- Hoja 4: Verificacion independiente vs Analisis_Capacidad de Madelyn
    ws = wb.create_sheet("Verificacion_vs_Madelyn")
    ws.append(["Celda", "Carga esperada SA101 (recalculada)", "Carga esperada SA102 (recalculada)",
               "HH esperadas SA101 (recalculada)", "HH esperadas SA102 (recalculada)",
               "Duracion total SA101 (recalculada)", "Duracion total SA102 (recalculada)"])
    for c in ws[1]:
        c.font = header_font; c.fill = header_fill
    for celda, r in resultados.items():
        ws.append([
            celda.replace("WorkCell", "Celda "),
            round(r["carga101"], 4), round(r["carga102"], 4),
            round(r["hh101"], 4), round(r["hh102"], 4),
            round(r["duracion_total101"], 4), round(r["duracion_total102"], 4),
        ])
    ws.append(["Nota:", "Estos valores se recalcularon desde 'Simio Aerospace Data.xlsx' de forma",
                "independiente al archivo de Madelyn, como control de calidad.", "", "", "", ""])
    ws.append(["Coinciden con Analisis_Capacidad.xlsx en todas las celdas salvo WorkCell5"])
    ws.append(["(ver nota de calidad en docs/00_contexto_y_plan.md sobre el Calculo 3 de esa hoja).", "", "", "", "", "", ""])
    for i, w in enumerate([14, 20, 20, 20, 20, 20, 20], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    wb.save(OUT_XLSX)
    print(f"Excel escrito en {OUT_XLSX}")

    # ---- Resumen en consola --------------------------------------------------
    print("\n=== RESUMEN ESCENARIO LOTE 75 AVIONES (45 SA101 / 30 SA102) ===")
    for celda in CELDAS:
        e = escenario[celda]
        print(f"{celda}: HH lote={e['hh_lote']:.1f}  ({e['pct_del_total_lote']:.1f}% del total)  "
              f"| dias solo M.O.={e['dias_solo_mano_obra']:.1f}  "
              f"| utilizacion bajo ciclo fijo={e['utilizacion_bajo_ciclo_fijo_pct']:.1f}%  "
              f"| pico mecanicos={e['pico_mecanicos_simultaneos']}/8")
    print(f"\nTotal HH del lote (5 celdas): {total_hh_lote:.1f}")
    print(f"Duracion de calendario del lote bajo ciclo fijo actual: {dias_calendario_lote} dias "
          f"(~{semanas_calendario_lote:.1f} semanas)")

    print(f"\n=== RIESGO DE TRABAJO DESPLAZADO (ventana de {CICLO_DIAS} dias = "
          f"{HORAS_EFECTIVAS_POR_CICLO} h efectivas) ===")
    for celda in CELDAS:
        e = escenario[celda]
        print(f"{celda}: SA101 dura {e['duracion_avion_sa101_h']:.1f} h "
              f"(desplazado {e['excedente_sa101_h']:.1f} h = {e['pct_trabajo_desplazado_sa101']:.0f}%)  "
              f"| SA102 dura {e['duracion_avion_sa102_h']:.1f} h "
              f"(desplazado {e['excedente_sa102_h']:.1f} h = {e['pct_trabajo_desplazado_sa102']:.0f}%)")


if __name__ == "__main__":
    main()
